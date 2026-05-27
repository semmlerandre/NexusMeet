from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, EmailStr, Field
from typing import List, Optional
from datetime import datetime, timedelta, timezone
import os
import logging
from pathlib import Path
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import shutil
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from jose import JWTError, jwt
from passlib.context import CryptContext
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Create uploads directory
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(UPLOADS_DIR)), name="uploads")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Auth setup
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
JWT_SECRET = os.environ.get("JWT_SECRET", "sua_chave_secreta")
JWT_ALGORITHM = os.environ.get("JWT_ALGORITHM", "HS256")
JWT_EXPIRATION_HOURS = int(os.environ.get("JWT_EXPIRATION_HOURS", "24"))

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Token inválido")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user or not user.get('ativo', True):
            raise HTTPException(status_code=401, detail="Usuário não encontrado")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Token inválido")

async def require_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get('perfil') != 'admin':
        raise HTTPException(status_code=403, detail="Acesso negado")
    return current_user

async def send_email_notification(to_email: str, subject: str, body_html: str):
    """Envia email se SMTP estiver configurado"""
    try:
        config = await db.system_config.find_one({"_id": "smtp_config"})
        if not config or not config.get('smtp_enabled'):
            logger.info(f"SMTP não configurado. Email não enviado para {to_email}")
            return False
        
        msg = MIMEMultipart()
        msg['From'] = config['smtp_from_email']
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body_html, 'html'))
        
        with smtplib.SMTP(config['smtp_server'], int(config['smtp_port'])) as server:
            server.starttls()
            server.login(config['smtp_user'], config['smtp_password'])
            server.send_message(msg)
        
        logger.info(f"Email enviado com sucesso para {to_email}")
        return True
    except Exception as e:
        logger.error(f"Erro ao enviar email: {str(e)}")
        return False

# Models
class UserCreate(BaseModel):
    nome: str
    email: EmailStr
    departamento: str
    senha: str
    perfil: Optional[str] = "user"

class UserLogin(BaseModel):
    email: EmailStr
    senha: str

class UserUpdate(BaseModel):
    relatorios_completos: bool

class RoomCreate(BaseModel):
    nome: str
    localizacao: str
    capacidade: int
    amenidades: Optional[List[str]] = []

class RoomUpdate(BaseModel):
    nome: Optional[str] = None
    localizacao: Optional[str] = None
    capacidade: Optional[int] = None
    ativa: Optional[bool] = None
    amenidades: Optional[List[str]] = None

class BookingCreate(BaseModel):
    room_id: int
    data_inicio: str
    data_fim: Optional[str] = None
    hora_inicio: str
    hora_fim: str

class SystemConfigUpdate(BaseModel):
    smtp_enabled: bool
    smtp_server: Optional[str] = None
    smtp_port: Optional[int] = None
    smtp_user: Optional[str] = None
    smtp_password: Optional[str] = None
    smtp_from_email: Optional[str] = None
    smtp_from_name: Optional[str] = None

class SystemThemeUpdate(BaseModel):
    system_name: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    accent_color: Optional[str] = None

class UserCreateManual(BaseModel):
    nome: str
    email: EmailStr
    departamento: str
    senha: str
    perfil: Optional[str] = "user"

class UserResetPassword(BaseModel):
    nova_senha: str

# Auth Routes
@api_router.post("/auth/register")
async def register_user(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    
    next_id = await db.counters.find_one_and_update(
        {"_id": "user_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    user_id = next_id["seq"] if next_id else 1
    
    user = {
        "id": user_id,
        "nome": user_data.nome,
        "email": user_data.email,
        "departamento": user_data.departamento,
        "senha_hash": hash_password(user_data.senha),
        "perfil": user_data.perfil,
        "ativo": True,
        "relatorios_completos": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    await db.logs.insert_one({"user_id": user_id, "acao": "REGISTRO", "timestamp": datetime.now(timezone.utc).isoformat()})
    
    user.pop('senha_hash')
    user.pop('_id', None)
    return user

@api_router.post("/auth/login")
async def login_user(user_data: UserLogin):
    user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if not user or not verify_password(user_data.senha, user['senha_hash']):
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    if not user.get('ativo', True):
        raise HTTPException(status_code=401, detail="Usuário inativo")
    
    token = create_access_token({"user_id": user['id'], "email": user['email'], "perfil": user['perfil']})
    await db.logs.insert_one({"user_id": user['id'], "acao": "LOGIN", "timestamp": datetime.now(timezone.utc).isoformat()})
    
    user.pop('senha_hash')
    return {"access_token": token, "token_type": "bearer", "user": user}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    current_user.pop('senha_hash', None)
    return current_user

# Users Routes
@api_router.get("/users")
async def list_users(current_user: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "senha_hash": 0}).to_list(1000)
    return users

@api_router.put("/users/{user_id}")
async def update_user(user_id: int, user_data: UserUpdate, current_user: dict = Depends(require_admin)):
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"relatorios_completos": user_data.relatorios_completos}}
    )
    await db.logs.insert_one({
        "user_id": current_user['id'],
        "acao": "ATUALIZAR_USUARIO",
        "detalhes": f"Usuário {user_id} - Relatórios completos: {user_data.relatorios_completos}",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    return {"message": "Usuário atualizado"}

@api_router.post("/users/bulk-import")
async def bulk_import_users(file: UploadFile = File(...), current_user: dict = Depends(require_admin)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Arquivo deve ser Excel")
    
    content = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(content))
    sheet = workbook.active
    
    total, criados, duplicados, erros = 0, 0, 0, []
    
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        total += 1
        if len(row) < 3 or not all([row[0], row[1], row[2]]):
            erros.append(f"Linha {idx}: Dados incompletos")
            continue
        
        nome, email, departamento = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip()
        
        if await db.users.find_one({"email": email}):
            duplicados += 1
            erros.append(f"Linha {idx}: Email {email} já existe")
            continue
        
        next_id = await db.counters.find_one_and_update(
            {"_id": "user_id"},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=True
        )
        user_id = next_id["seq"] if next_id else 1
        
        await db.users.insert_one({
            "id": user_id,
            "nome": nome,
            "email": email,
            "departamento": departamento,
            "senha_hash": hash_password("Sala@123"),
            "perfil": "user",
            "ativo": True,
            "relatorios_completos": False,
            "primeiro_acesso": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        criados += 1
    
    await db.logs.insert_one({
        "user_id": current_user['id'],
        "acao": "IMPORTACAO_USUARIOS",
        "detalhes": f"{criados} criados",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    
    return {"total_processados": total, "usuarios_criados": criados, "usuarios_duplicados": duplicados, "erros": erros}



# Criar usuário manualmente (Admin)
@api_router.post("/users/create-manual")
async def create_user_manual(user_data: UserCreateManual, current_user: dict = Depends(require_admin)):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email já cadastrado")
    
    next_id = await db.counters.find_one_and_update({"_id": "user_id"}, {"$inc": {"seq": 1}}, upsert=True, return_document=True)
    user_id = next_id["seq"] if next_id else 1
    
    new_user = {
        "id": user_id,
        "nome": user_data.nome,
        "email": user_data.email,
        "departamento": user_data.departamento,
        "senha_hash": hash_password(user_data.senha),
        "perfil": user_data.perfil,
        "ativo": True,
        "relatorios_completos": False,
        "primeiro_acesso": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(new_user)
    await db.logs.insert_one({"user_id": current_user['id'], "acao": "CRIAR_USUARIO", "detalhes": f"Usuário {new_user['nome']} criado", "timestamp": datetime.now(timezone.utc).isoformat()})
    
    new_user.pop('_id')
    new_user.pop('senha_hash')
    return new_user

# Resetar senha (Admin)
@api_router.post("/users/{user_id}/reset-password")
async def reset_user_password(user_id: int, password_data: UserResetPassword, current_user: dict = Depends(require_admin)):
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    await db.users.update_one({"id": user_id}, {"$set": {"senha_hash": hash_password(password_data.nova_senha), "primeiro_acesso": True}})
    await db.logs.insert_one({"user_id": current_user['id'], "acao": "RESETAR_SENHA", "detalhes": f"Senha resetada: {user['nome']}", "timestamp": datetime.now(timezone.utc).isoformat()})
    
    return {"message": "Senha resetada. Usuário deverá alterar no próximo login."}

# Alterar própria senha
@api_router.post("/auth/change-password")
async def change_password(current_password: str, new_password: str, current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user['id']})
    if not verify_password(current_password, user['senha_hash']):
        raise HTTPException(status_code=400, detail="Senha atual incorreta")
    
    await db.users.update_one({"id": current_user['id']}, {"$set": {"senha_hash": hash_password(new_password), "primeiro_acesso": False}})
    await db.logs.insert_one({"user_id": current_user['id'], "acao": "ALTERAR_SENHA", "timestamp": datetime.now(timezone.utc).isoformat()})
    
    return {"message": "Senha alterada com sucesso"}

# Rooms Routes
@api_router.post("/rooms")
async def create_room(room_data: RoomCreate, current_user: dict = Depends(require_admin)):
    next_id = await db.counters.find_one_and_update(
        {"_id": "room_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    room_id = next_id["seq"] if next_id else 1
    
    room = {
        "id": room_id,
        "nome": room_data.nome,
        "localizacao": room_data.localizacao,
        "capacidade": room_data.capacidade,
        "amenidades": room_data.amenidades or [],
        "ativa": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.rooms.insert_one(room)
    await db.logs.insert_one({
        "user_id": current_user['id'],
        "acao": "CRIAR_SALA",
        "detalhes": f"Sala {room['nome']} criada",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    room.pop('_id')
    return room

@api_router.get("/rooms")
async def list_rooms(current_user: dict = Depends(get_current_user)):
    rooms = await db.rooms.find({"ativa": True}, {"_id": 0}).to_list(1000)
    return rooms

@api_router.get("/rooms/{room_id}")
async def get_room(room_id: int, current_user: dict = Depends(get_current_user)):
    room = await db.rooms.find_one({"id": room_id}, {"_id": 0})
    if not room:
        raise HTTPException(status_code=404, detail="Sala não encontrada")
    return room

@api_router.put("/rooms/{room_id}")
async def update_room(room_id: int, room_data: RoomUpdate, current_user: dict = Depends(require_admin)):
    room = await db.rooms.find_one({"id": room_id})
    if not room:
        raise HTTPException(status_code=404, detail="Sala não encontrada")
    
    update_data = {k: v for k, v in room_data.model_dump(exclude_unset=True).items() if v is not None}
    
    if update_data:
        await db.rooms.update_one({"id": room_id}, {"$set": update_data})
        await db.logs.insert_one({
            "user_id": current_user['id'],
            "acao": "ATUALIZAR_SALA",
            "detalhes": f"Sala ID {room_id} atualizada",
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
    
    updated_room = await db.rooms.find_one({"id": room_id}, {"_id": 0})
    return updated_room

@api_router.delete("/rooms/{room_id}")
async def delete_room(room_id: int, current_user: dict = Depends(require_admin)):
    await db.rooms.update_one({"id": room_id}, {"$set": {"ativa": False}})
    await db.logs.insert_one({
        "user_id": current_user['id'],
        "acao": "DELETAR_SALA",
        "detalhes": f"Sala ID {room_id} desativada",
        "timestamp": datetime.now(timezone.utc).isoformat()
    })
    return {"message": "Sala desativada"}

# Bookings Routes
@api_router.post("/bookings")
async def create_booking(booking_data: BookingCreate, current_user: dict = Depends(get_current_user)):
    room = await db.rooms.find_one({"id": booking_data.room_id, "ativa": True}, {"_id": 0})
    if not room:
        raise HTTPException(status_code=404, detail="Sala não encontrada")
    
    data_fim = booking_data.data_fim if booking_data.data_fim else booking_data.data_inicio
    data_inicio_obj = datetime.strptime(booking_data.data_inicio, '%Y-%m-%d')
    data_fim_obj = datetime.strptime(data_fim, '%Y-%m-%d')
    
    if data_fim_obj < data_inicio_obj:
        raise HTTPException(status_code=400, detail="Data fim deve ser maior ou igual à data início")
    
    # Gerar lista de datas
    datas = []
    current_date = data_inicio_obj
    while current_date <= data_fim_obj:
        datas.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)
    
    # Verificar disponibilidade para todas as datas
    for data in datas:
        existing = await db.bookings.find({
            "room_id": booking_data.room_id,
            "data": data,
            "status": "confirmed"
        }, {"_id": 0}).to_list(1000)
        
        for b in existing:
            if not (booking_data.hora_fim <= b['hora_inicio'] or booking_data.hora_inicio >= b['hora_fim']):
                raise HTTPException(status_code=400, detail=f"Sala não disponível em {data} no horário solicitado")
    
    # Criar reservas para cada data
    bookings_criadas = []
    for data in datas:
        next_id = await db.counters.find_one_and_update(
            {"_id": "booking_id"},
            {"$inc": {"seq": 1}},
            upsert=True,
            return_document=True
        )
        booking_id = next_id["seq"] if next_id else 1
        
        booking = {
            "id": booking_id,
            "room_id": booking_data.room_id,
            "user_id": current_user['id'],
            "data": data,
            "hora_inicio": booking_data.hora_inicio,
            "hora_fim": booking_data.hora_fim,
            "status": "confirmed",
            "room_name": room['nome'],
            "user_name": current_user['nome'],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.bookings.insert_one(booking)
        booking.pop('_id')
        bookings_criadas.append(booking)
    
    await db.logs.insert_one({"user_id": current_user['id'], "acao": "CRIAR_RESERVA", "detalhes": f"{len(datas)} dias", "timestamp": datetime.now(timezone.utc).isoformat()})
    
    # Enviar email
    if len(datas) == 1:
        subject = "Reserva Confirmada - NexusMeet"
        body = f"""
        <h2>Reserva Confirmada</h2>
        <p>Sua reserva foi confirmada com sucesso!</p>
        <p><strong>Sala:</strong> {room['nome']}</p>
        <p><strong>Data:</strong> {datas[0]}</p>
        <p><strong>Horário:</strong> {booking_data.hora_inicio} - {booking_data.hora_fim}</p>
        """
    else:
        subject = "Reservas Confirmadas - NexusMeet"
        body = f"""
        <h2>Reservas Confirmadas</h2>
        <p>Suas reservas foram confirmadas com sucesso!</p>
        <p><strong>Sala:</strong> {room['nome']}</p>
        <p><strong>Período:</strong> {datas[0]} a {datas[-1]} ({len(datas)} dias)</p>
        <p><strong>Horário:</strong> {booking_data.hora_inicio} - {booking_data.hora_fim}</p>
        """
    
    await send_email_notification(current_user['email'], subject, body)
    
    return {"bookings": bookings_criadas, "total": len(bookings_criadas)}

@api_router.get("/bookings")
async def list_bookings(current_user: dict = Depends(get_current_user)):
    bookings = await db.bookings.find({"status": "confirmed"}, {"_id": 0}).to_list(1000)
    return bookings

@api_router.get("/bookings/my")
async def list_my_bookings(current_user: dict = Depends(get_current_user)):
    bookings = await db.bookings.find({"user_id": current_user['id'], "status": "confirmed"}, {"_id": 0}).to_list(1000)
    return bookings

@api_router.delete("/bookings/{booking_id}")
async def cancel_booking(booking_id: int, current_user: dict = Depends(get_current_user)):
    booking = await db.bookings.find_one({"id": booking_id}, {"_id": 0})
    if not booking:
        raise HTTPException(status_code=404, detail="Reserva não encontrada")
    if booking['user_id'] != current_user['id'] and current_user['perfil'] != 'admin':
        raise HTTPException(status_code=403, detail="Sem permissão")
    
    await db.bookings.update_one({"id": booking_id}, {"$set": {"status": "cancelled"}})
    await db.logs.insert_one({"user_id": current_user['id'], "acao": "CANCELAR_RESERVA", "timestamp": datetime.now(timezone.utc).isoformat()})
    
    # Enviar email
    user = await db.users.find_one({"id": booking['user_id']})
    if user:
        subject = "Reserva Cancelada - NexusMeet"
        body = f"""
        <h2>Reserva Cancelada</h2>
        <p>Sua reserva foi cancelada.</p>
        <p><strong>Sala:</strong> {booking.get('room_name', 'N/A')}</p>
        <p><strong>Data:</strong> {booking['data']}</p>
        <p><strong>Horário:</strong> {booking['hora_inicio']} - {booking['hora_fim']}</p>
        """
        await send_email_notification(user['email'], subject, body)
    
    return {"message": "Reserva cancelada"}

# Reports Routes
@api_router.get("/reports/bookings")
async def generate_report(start_date: Optional[str] = None, end_date: Optional[str] = None, 
                         user_id: Optional[int] = None, room_id: Optional[int] = None,
                         current_user: dict = Depends(get_current_user)):
    # Verificar permissões
    pode_ver_todos = current_user['perfil'] == 'admin' or current_user.get('relatorios_completos', False)
    
    query = {"status": "confirmed"}
    
    if not pode_ver_todos:
        query["user_id"] = current_user['id']
    elif user_id:
        query["user_id"] = user_id
    
    if room_id:
        query["room_id"] = room_id
    
    if start_date:
        query["data"] = {"$gte": start_date}
    if end_date:
        if "data" in query:
            query["data"]["$lte"] = end_date
        else:
            query["data"] = {"$lte": end_date}
    
    bookings = await db.bookings.find(query, {"_id": 0}).sort("data", -1).to_list(10000)
    
    return {
        "total_reservas": len(bookings),
        "reservas": bookings,
        "pode_ver_todos": pode_ver_todos
    }

@api_router.get("/reports/export")
async def export_report(start_date: Optional[str] = None, end_date: Optional[str] = None,
                       user_id: Optional[int] = None, room_id: Optional[int] = None,
                       current_user: dict = Depends(get_current_user)):
    # Mesma lógica de permissões
    pode_ver_todos = current_user['perfil'] == 'admin' or current_user.get('relatorios_completos', False)
    
    query = {"status": "confirmed"}
    if not pode_ver_todos:
        query["user_id"] = current_user['id']
    elif user_id:
        query["user_id"] = user_id
    if room_id:
        query["room_id"] = room_id
    if start_date:
        query["data"] = {"$gte": start_date}
    if end_date:
        if "data" in query:
            query["data"]["$lte"] = end_date
        else:
            query["data"] = {"$lte": end_date}
    
    bookings = await db.bookings.find(query, {"_id": 0}).sort("data", 1).to_list(10000)
    
    # Criar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Reservas"
    
    # Cabeçalhos
    headers = ['ID', 'Sala', 'Usuário', 'Data', 'Hora Início', 'Hora Fim', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Dados
    for row_idx, booking in enumerate(bookings, start=2):
        ws.cell(row=row_idx, column=1, value=booking['id'])
        ws.cell(row=row_idx, column=2, value=booking.get('room_name', 'N/A'))
        ws.cell(row=row_idx, column=3, value=booking.get('user_name', 'N/A'))
        ws.cell(row=row_idx, column=4, value=booking['data'])
        ws.cell(row=row_idx, column=5, value=booking['hora_inicio'])
        ws.cell(row=row_idx, column=6, value=booking['hora_fim'])
        ws.cell(row=row_idx, column=7, value=booking['status'])
    
    # Ajustar largura das colunas
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 20
    
    # Salvar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=relatorio_reservas.xlsx"}
    )

# System Config Routes
@api_router.get("/config")
async def get_config():
    config = await db.system_config.find_one({"_id": "system_config"}, {"_id": 0, "smtp_password": 0})
    if not config:
        config = {
            "logo_url": None,
            "login_bg_url": None
        }
    return config


# Configuração de Tema
@api_router.get("/config/theme")
async def get_theme_config():
    config = await db.system_config.find_one({"_id": "theme_config"}, {"_id": 0})
    if not config:
        config = {
            "system_name": "NexusMeet",
            "primary_color": "#1e293b",
            "secondary_color": "#f1f5f9",
            "accent_color": "#3b82f6"
        }
    return config

@api_router.put("/config/theme")
async def update_theme_config(theme_data: SystemThemeUpdate, current_user: dict = Depends(require_admin)):
    theme_dict = {k: v for k, v in theme_data.model_dump().items() if v is not None}
    
    if theme_dict:
        await db.system_config.update_one(
            {"_id": "theme_config"},
            {"$set": theme_dict},
            upsert=True
        )
        await db.logs.insert_one({
            "user_id": current_user['id'],
            "acao": "ATUALIZAR_TEMA",
            "detalhes": f"Tema atualizado: {list(theme_dict.keys())}",
            "timestamp": datetime.now(timezone.utc).isoformat()
        })
    
    return {"message": "Tema atualizado com sucesso"}


@api_router.get("/config/smtp")
async def get_smtp_config(current_user: dict = Depends(require_admin)):
    config = await db.system_config.find_one({"_id": "smtp_config"}, {"_id": 0, "smtp_password": 0})
    return config or {}

@api_router.put("/config/smtp")
async def update_smtp_config(config_data: SystemConfigUpdate, current_user: dict = Depends(require_admin)):
    config_dict = config_data.model_dump()
    await db.system_config.update_one(
        {"_id": "smtp_config"},
        {"$set": config_dict},
        upsert=True
    )
    await db.logs.insert_one({"user_id": current_user['id'], "acao": "ATUALIZAR_SMTP", "timestamp": datetime.now(timezone.utc).isoformat()})
    return {"message": "Configuração SMTP atualizada"}

@api_router.post("/config/smtp/test")
async def test_smtp(email: str, current_user: dict = Depends(require_admin)):
    success = await send_email_notification(
        email,
        "Teste de Configuração SMTP - NexusMeet",
        "<h2>Teste realizado com sucesso!</h2><p>Sua configuração SMTP está funcionando corretamente.</p>"
    )
    if success:
        return {"message": "Email de teste enviado com sucesso"}
    else:
        raise HTTPException(status_code=500, detail="Erro ao enviar email de teste")

@api_router.post("/config/upload/{type}")
async def upload_file(type: str, file: UploadFile = File(...), current_user: dict = Depends(require_admin)):
    if type not in ['logo', 'login_bg']:
        raise HTTPException(status_code=400, detail="Tipo inválido")
    
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Arquivo deve ser uma imagem")
    
    # Salvar arquivo
    ext = file.filename.split('.')[-1]
    filename = f"{type}.{ext}"
    filepath = UPLOADS_DIR / filename
    
    with open(filepath, "wb") as f:
        content = await file.read()
        f.write(content)
    
    url = f"/uploads/{filename}"
    
    # Atualizar config
    field = "logo_url" if type == "logo" else "login_bg_url"
    await db.system_config.update_one(
        {"_id": "system_config"},
        {"$set": {field: url}},
        upsert=True
    )
    
    await db.logs.insert_one({"user_id": current_user['id'], "acao": f"UPLOAD_{type.upper()}", "timestamp": datetime.now(timezone.utc).isoformat()})
    
    return {"url": url}

@api_router.get("/logs")
async def list_logs(current_user: dict = Depends(require_admin)):
    logs = await db.logs.find({}, {"_id": 0}).sort("timestamp", -1).limit(100).to_list(100)
    
    # Adicionar nomes dos usuários
    result = []
    for log in logs:
        if log.get('user_id'):
            user = await db.users.find_one({"id": log['user_id']}, {"_id": 0, "nome": 1})
            log['user_name'] = user['nome'] if user else 'Desconhecido'
        result.append(log)
    
    return result

# Dashboard Stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    # Total de salas
    total_rooms = await db.rooms.count_documents({"ativa": True})
    
    # Reservas de hoje
    bookings_today = await db.bookings.count_documents({
        "data": today,
        "status": "confirmed"
    })
    
    # Minhas reservas ativas
    my_bookings = await db.bookings.count_documents({
        "user_id": current_user['id'],
        "status": "confirmed",
        "data": {"$gte": today}
    })
    
    # Total de usuários (só admin vê)
    total_users = 0
    if current_user['perfil'] == 'admin':
        total_users = await db.users.count_documents({})
    
    # Próximas reservas (próximos 7 dias)
    future_date = (datetime.now(timezone.utc) + timedelta(days=7)).strftime('%Y-%m-%d')
    upcoming_bookings = await db.bookings.find({
        "data": {"$gte": today, "$lte": future_date},
        "status": "confirmed"
    }, {"_id": 0}).sort("data", 1).limit(5).to_list(5)
    
    # Status das salas (ocupadas agora)
    current_time = datetime.now(timezone.utc).strftime('%H:%M')
    rooms_occupied = await db.bookings.count_documents({
        "data": today,
        "status": "confirmed",
        "hora_inicio": {"$lte": current_time},
        "hora_fim": {"$gte": current_time}
    })
    
    return {
        "total_rooms": total_rooms,
        "rooms_available": total_rooms - rooms_occupied,
        "rooms_occupied": rooms_occupied,
        "bookings_today": bookings_today,
        "my_bookings": my_bookings,
        "total_users": total_users,
        "upcoming_bookings": upcoming_bookings
    }

# Room Schedule for Today
@api_router.get("/rooms/{room_id}/schedule")
async def get_room_schedule(room_id: int, date: str = None, current_user: dict = Depends(get_current_user)):
    if not date:
        date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    room = await db.rooms.find_one({"id": room_id}, {"_id": 0})
    if not room:
        raise HTTPException(status_code=404, detail="Sala não encontrada")
    
    bookings = await db.bookings.find({
        "room_id": room_id,
        "data": date,
        "status": "confirmed"
    }, {"_id": 0}).sort("hora_inicio", 1).to_list(1000)
    
    return {
        "room": room,
        "date": date,
        "bookings": bookings
    }

# Download deploy script
@api_router.get("/deploy/script")
async def get_deploy_script(current_user: dict = Depends(require_admin)):
    script_path = Path("/app/scripts/deploy-auto-port.sh")
    if not script_path.exists():
        raise HTTPException(status_code=404, detail="Script não encontrado")
    
    from fastapi.responses import FileResponse
    return FileResponse(
        path=script_path,
        filename="deploy-nexusmeet.sh",
        media_type="application/x-sh"
    )

@api_router.get("/")
async def root():
    return {"message": "Sistema de Agendamento de Salas - API v1.0"}

app.include_router(api_router)

@app.on_event("startup")
async def startup_event():
    logger.info("Sistema iniciado")
    admin = await db.users.find_one({"email": "admin@sistema.com"})
    if not admin:
        await db.users.insert_one({
            "id": 1,
            "nome": "Administrador",
            "email": "admin@sistema.com",
            "departamento": "TI",
            "senha_hash": hash_password("admin123"),
            "perfil": "admin",
            "ativo": True,
            "relatorios_completos": True,
            "primeiro_acesso": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        await db.counters.update_one({"_id": "user_id"}, {"$set": {"seq": 1}}, upsert=True)
        logger.info("Admin criado: admin@sistema.com / admin123")
    
    # Criar salas de exemplo
    count = await db.rooms.count_documents({})
    if count == 0:
        salas = [
            {"id": 1, "nome": "Sala Executiva", "localizacao": "3º Andar - Ala A", "capacidade": 12, "ativa": True},
            {"id": 2, "nome": "Sala de Brainstorm", "localizacao": "2º Andar - Ala B", "capacidade": 8, "ativa": True},
            {"id": 3, "nome": "Sala de Videoconferência", "localizacao": "1º Andar", "capacidade": 6, "ativa": True},
            {"id": 4, "nome": "Auditório Principal", "localizacao": "Térreo", "capacidade": 50, "ativa": True}
        ]
        await db.rooms.insert_many(salas)
        await db.counters.update_one({"_id": "room_id"}, {"$set": {"seq": 4}}, upsert=True)
        logger.info("4 salas de exemplo criadas")

@app.on_event("shutdown")
async def shutdown_event():
    client.close()